#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
etl_montature.py
=================
Ricostruisce automaticamente il file "Base MONTATURE" a partire dai file
esportati quotidianamente (giacenza, vendite, listini, movimenti, sottoscorta),
replicando la logica di join che prima veniva fatta a mano in Access.

CHIAVE DI JOIN: "codice a barre" + "codice filiale" (es. "41031001" + "G" =
ID_GIACENZA "41031001G"), esattamente come nel file Access originale.

USO:
    python3 etl_montature.py --input-dir "/percorso/Export FOCUS DEPOSITO" \
                              --ref-dir "/percorso/tabelle_riferimento" \
                              --output "Base MONTATURE.xlsx"

I file sorgente (percorso Dropbox assoluto, vedi dizionario FILES qui sotto):
    /Export FOCUS DEPOSITO/BASE/Listini MONTATURE 2017.txt
    /Export FOCUS DEPOSITO/BASE/Sottoscorta MONTATURE.txt
    /Export FOCUS DEPOSITO/BASE/Movimenti ACQUISTO.txt
    /Export FOCUS DEPOSITO/RICERCHE BASE/GIACENZA MONTATURE al 2017.txt
    /Power BI/Dati/VENDUTO/Vendita TUTTI MAGAZZINI.xlsx   (vendite multi-anno,
                                                            solo righe "VEN")

Nota: dal 14/07/2026 i file vivono su un Dropbox diverso da quello originale
(un altro PC, aggiornato ogni notte) e le vendite arrivano da un unico export
Power BI con lo storico di piu' anni, non piu' da due file separati per
anno corrente/precedente.

--input-dir deve contenere, per ogni voce di FILES, un file con il nome
indicato in "local_name" (es. giacenza.txt, vendite.xlsx...). E' app.py che si
occupa di scaricare da Dropbox ogni file al suo "local_name" dentro una
cartella temporanea prima di chiamare build().

Le tabelle di riferimento "manuali" (colonne che in Access erano compilate a
mano e non provengono da nessuno dei file quotidiani) vengono lette da
--ref-dir e AGGIORNATE in automatico con i nuovi articoli trovati (lasciati
vuoti, pronti per essere compilati):
    ref_fornitore2.csv        fornitore -> Fornitore 2
    ref_articolo_manuale.csv  Codice a Barre -> Top, POSIZIONE GRIGLIA,
                               Occhiali con CLIP, Personale, scorta minima

"Marchi Attivi" (colonna AG) NON e' piu' una tabella manuale: dal 17/07/2026
si calcola automaticamente confrontando "Marchio" con due elenchi di marchi
attivi, distinti per "Tipo Lenti" (vedi ref-dir):
    ref_marchi_attivi_sole.csv   elenco marchi attivi per Tipo Lenti = sole
    ref_marchi_attivi_vista.csv  elenco marchi attivi per Tipo Lenti = vista
Se "Tipo Lenti" e' "sole" e il marchio e' nell'elenco sole (o "vista" e
nell'elenco vista) -> "SI", altrimenti vuoto (comprese tutte le altre
categorie di lenti: predisposti, clip-on, protettivi, nuoto, ecc., che non
sono coperte da questi elenchi e restano sempre vuote). Il confronto ignora
maiuscole/minuscole e spazi iniziali/finali. Per aggiornare gli elenchi
attivi, sostituire semplicemente questi due file CSV (stessa struttura: una
colonna "Marchio").

NOTE / IPOTESI (confermate con Salvo il 14/07/2026):
  - Vengono escluse le combinazioni articolo+negozio con giacenza a ZERO e
    NESSUNA vendita negli ultimi due anni solari (anno corrente + anno
    precedente). Es. nel 2026: fuori chi ha giacenza 0 e zero vendite sia nel
    2026 sia nel 2025.
  - Dal file vendite si conta solo "Tipo operazione" = VEN (non ING).
  - "Quantita Disponibile" = quantita magazzino - quantita prenotata (min 0)
  - "Vendita {ANNO-1} CUMULATO" = vendite anno precedente dal 1/1 fino allo
    stesso giorno/mese di oggi (confronto omogeneo anno su anno)
  - "Vendita {ANNO-2}" ora e' calcolata con dati reali (il nuovo file vendite
    copre piu' anni), prima restava sempre vuota
  - "Filiale BIS" = uguale a "Filiale"

REGOLE FISSE (confermate con Salvo il 17/07/2026):
  - "Fornitore 2" (colonna B) = "KERING EYEWEAR" quando Marchio e'
    esattamente "CARTIER" (non da tabella manuale, sovrascrive la tabella).
  - Quando Marchio e' esattamente "Ray-Ban", dal "modello" vengono tolte le
    parole SOLE, VISTA, SUN, OPTICAL (senza lasciare lo spazio prima), PRIMA
    di calcolare le colonne G/H/I ("Modello + Colore + Calibro..."). Altri
    marchi simili (es. "Ray-Ban Junior", "Ray-ban Meta") non sono toccati.
  - "data ult acquisto" (colonna AA) = presa direttamente dal campo
    "data ult. acquisto" della GIACENZA (non piu' calcolata da MOVIMENTI).
  - "data" (colonna AB) = calcolata da MOVIMENTI ACQUISTO.txt come data piu'
    recente (tipo operazione = ACQ) per lo stesso codice a barre.
  - "Prezzo Di Acquisto Scheda Scontato" (colonna AJ) e "Prezzo Di Vendita
    Scheda Scontato" (colonna AK) = presi direttamente dai campi "prezzo
    acquisto" e "prezzo di vendita" della GIACENZA (non piu' dalla
    sottoscorta). Sconto Acquisto (AN), Sconto Vendita (AO) e Fattore di
    RICARICO (AP), che usano questi due valori come base, seguono di
    conseguenza gli stessi nuovi prezzi.
  - "Occhiali con CLIP" (colonna AH) NON e' piu' una tabella manuale:
    "SI" solo se Tipo Lenti = VISTA e la colonna G ("Modello + Colore +
    Calibro + CAT", gia' costruita a quel punto) contiene la parola CLIP
    (anche insieme ad altre parole, es. "CLIP-ON"), altrimenti vuoto.
  - Sconto Acquisto (AN) e Sconto Vendita (AO) = valori interi arrotondati
    (nessun decimale). Fattore di RICARICO (AP) = arrotondato a un decimale.
"""
import argparse
import csv
import datetime as dt
import os
import re
import sys
from collections import defaultdict

# ---------------------------------------------------------------------------
# CONFIGURAZIONE
# ---------------------------------------------------------------------------

# Ogni file sorgente ha un percorso Dropbox assoluto (da settembre 2026 non
# sono piu' tutti nella stessa cartella: la giacenza/listini/sottoscorta/
# movimenti restano in "Export FOCUS DEPOSITO", le vendite arrivano invece da
# un unico file esportato da Power BI con lo storico di piu' anni).
FILES = {
    "giacenza": {
        "dropbox_path": "/Export FOCUS DEPOSITO/RICERCHE BASE/GIACENZA MONTATURE al 2017.txt",
        "local_name": "giacenza.txt",
        "format": "tsv",
    },
    "listini": {
        "dropbox_path": "/Export FOCUS DEPOSITO/BASE/Listini MONTATURE 2017.txt",
        "local_name": "listini.txt",
        "format": "tsv",
    },
    "sottoscorta": {
        "dropbox_path": "/Export FOCUS DEPOSITO/BASE/Sottoscorta MONTATURE.txt",
        "local_name": "sottoscorta.txt",
        "format": "tsv",
    },
    "movimenti": {
        "dropbox_path": "/Export FOCUS DEPOSITO/BASE/Movimenti ACQUISTO.txt",
        "local_name": "movimenti.txt",
        "format": "tsv",
    },
    "vendite": {
        "dropbox_path": "/Power BI/Dati/VENDUTO/Vendita TUTTI MAGAZZINI.xlsx",
        "local_name": "vendite.xlsx",
        "format": "xlsx",
    },
}

# Elenco filiali "vere" da tenere nel file finale (codice -> nome).
# Estratto dal file Access esistente. Se aprono un nuovo punto vendita,
# aggiungere qui la riga corrispondente.
FILIALI = {
    "A": ".Deposito",
    "A2": "Catania",
    "B": "Paterno",
    "C": "Acireale",
    "D": "Catania",
    "E": "Augusta",
    "G": "Giarre",
    "H": "Lentini",
    "I": "Portali",
    "IR": "Portali",
    "L": "Misterbianco",
    "R": "Paolo",
    "UR": "Paolo Bis",
    "V": "OUTLET",
    "Y": "Portali",
    "Z": "Viale Africa",
}

OUTPUT_COLUMNS = [
    "fornitore", "Fornitore 2", "Marchio", "modello", "Modello CARTIER", "SKU",
    "Modello + Colore + Calibro + CAT", "Modello + Colore + Calibro", "Colore + Calibro",
    "colore", "colore 2", "calibro", "ponte", "materiale", "utente", "Tipo Lenti",
    "Filiale", "Filiale BIS", "quantita magazzino", "Quantita Disponibile",
    "quantita in arrivo", "scorta minima",
    None, None, None, None,  # colonne vendite per anno, nomi dinamici (vedi build)
    "data ult acquisto", "data", "data ult vendita", "Categoria Filtro",
    "POSIZIONE GRIGLIA", "Top", "Marchi Attivi", "Occhiali con CLIP", "Personale",
    "Prezzo Di Acquisto Scheda Scontato", "Prezzo Di Vendita Scheda Scontato",
    "Prezzo Acquisto Listino Intero", "Prezzo Vendita Listino Intero",
    "Sconto Acquisto", "Sconto Vendita", "Fattore di RICARICO (su prezzi scontati)",
    "Codice a Barre", "ID_GIACENZA", "Filiale Focus",
]

# ---------------------------------------------------------------------------
# UTILITY
# ---------------------------------------------------------------------------


def parse_it_number(s):
    """Converte '143,65' -> 143.65. Stringa vuota -> None."""
    if s is None:
        return None
    s = s.strip()
    if s == "":
        return None
    s = s.replace(".", "").replace(",", ".") if "," in s else s
    try:
        return float(s)
    except ValueError:
        return None


def parse_it_date(s):
    """Converte 'gg/mm/aaaa' -> datetime.date. Stringa vuota -> None."""
    if not s:
        return None
    s = s.strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def load_tsv(path):
    """Legge un file .txt esportato (tab-separated, virgolette, encoding latin-1).
    Usato solo per i file piccoli (vendite). Per i file grandi (giacenza,
    listini, sottoscorta, movimenti) si usa iter_tsv per non saturare la RAM."""
    with open(path, encoding="latin-1", newline="") as f:
        reader = csv.reader(f, delimiter="\t", quotechar='"')
        header = next(reader)
        rows = [dict(zip(header, row)) for row in reader]
    return rows


def iter_tsv(path):
    """Ritorna (indice_colonne, reader, filehandle) senza caricare tutto il
    file in memoria. Ogni riga e' una lista posizionale."""
    f = open(path, encoding="latin-1", newline="")
    reader = csv.reader(f, delimiter="\t", quotechar='"')
    header = next(reader)
    idx = {name: i for i, name in enumerate(header)}
    return idx, reader, f


def load_ref_csv(path, key_field):
    d = {}
    if os.path.exists(path):
        with open(path, encoding="utf-8", newline="") as f:
            for row in csv.DictReader(f):
                d[row[key_field]] = row
    return d


def save_ref_csv(path, fieldnames, key_field, data):
    rows = sorted(data.values(), key=lambda r: str(r.get(key_field, "")))
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in fieldnames})


def load_brand_set(path):
    """Legge un CSV a una colonna 'Marchio' e ritorna un set di marchi
    normalizzati (maiuscolo, senza spazi iniziali/finali). Usato per gli
    elenchi 'Marchi Attivi' (sole/vista)."""
    brands = set()
    if os.path.exists(path):
        with open(path, encoding="utf-8", newline="") as f:
            for row in csv.DictReader(f):
                v = (row.get("Marchio") or "").strip()
                if v:
                    brands.add(v.upper())
    return brands


# Parole da rimuovere dal "modello" quando il Marchio e' esattamente Ray-Ban
# (regola fissa, non da tabella manuale). Si rimuove anche lo spazio che le
# precede, ovunque compaiano nella stringa (di solito in coda, es. "2001 SOLE").
_RAYBAN_STRIP_WORDS = ("SOLE", "VISTA", "SUN", "OPTICAL")
_RAYBAN_STRIP_RE = re.compile(r"\s*\b(?:" + "|".join(_RAYBAN_STRIP_WORDS) + r")\b", re.IGNORECASE)


def clean_modello_rayban(marchio, modello):
    """Se il Marchio e' esattamente 'Ray-Ban', toglie da 'modello' le parole
    SOLE, VISTA, SUN, OPTICAL (senza lasciare lo spazio prima). Per gli altri
    marchi (es. 'Ray-Ban Junior', 'Ray-ban Meta') il modello resta invariato."""
    if not modello:
        return modello
    if (marchio or "").strip() != "Ray-Ban":
        return modello
    cleaned = _RAYBAN_STRIP_RE.sub("", modello)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned


def get_occhiali_clip(tipo_lenti, colonna_g):
    """Colonna AH ('Occhiali con CLIP'): 'SI' solo se Tipo Lenti = VISTA e la
    colonna G ('Modello + Colore + Calibro + CAT', gia' costruita) contiene
    la parola CLIP, anche insieme ad altre parole (es. 'CLIP-ON'). Calcolata
    DOPO aver costruito la colonna G."""
    if (tipo_lenti or "").strip().upper() != "VISTA":
        return ""
    if "CLIP" in (colonna_g or "").upper():
        return "SI"
    return ""


# ---------------------------------------------------------------------------
# COSTRUZIONE DEL FILE
# ---------------------------------------------------------------------------


def get(row, idx, name):
    i = idx.get(name)
    return row[i] if i is not None and i < len(row) else ""


def build(input_dir, ref_dir, today=None):
    today = today or dt.date.today()
    year = today.year

    def p(key):
        return os.path.join(input_dir, FILES[key]["local_name"])

    print("Lettura GIACENZA (file guida)...", file=sys.stderr)
    idx_g, reader_g, fh_g = iter_tsv(p("giacenza"))
    giac_rows = []
    barcodes_rilevanti = set()
    for row in reader_g:
        fil_code = row[idx_g["filiale"]]
        if fil_code not in FILIALI:
            continue
        giac_rows.append(row)
        barcodes_rilevanti.add(row[idx_g["codice a barre"]])
    fh_g.close()
    print(f"  righe giacenza rilevanti: {len(giac_rows)}, barcode distinti: {len(barcodes_rilevanti)}", file=sys.stderr)

    print("Lettura LISTINI (solo barcode rilevanti)...", file=sys.stderr)
    listini_by_barcode = {}
    idx_l, reader_l, fh_l = iter_tsv(p("listini"))
    for row in reader_l:
        bc = row[idx_l["codice a barre"]]
        if bc in barcodes_rilevanti:
            listini_by_barcode[bc] = {
                "SKU": get(row, idx_l, "SKU"),
                "colore 2": get(row, idx_l, "colore 2"),
                "ponte": get(row, idx_l, "ponte"),
                "prezzo acquisto": get(row, idx_l, "prezzo acquisto"),
                "prezzo di vendita": get(row, idx_l, "prezzo di vendita"),
            }
    fh_l.close()

    print("Lettura SOTTOSCORTA (solo barcode rilevanti)...", file=sys.stderr)
    sottoscorta_by_barcode = {}
    idx_s, reader_s, fh_s = iter_tsv(p("sottoscorta"))
    for row in reader_s:
        bc = row[idx_s["codice a barre"]]
        if bc in barcodes_rilevanti:
            sottoscorta_by_barcode[bc] = {
                "SKU": get(row, idx_s, "SKU"),
                "colore 2": get(row, idx_s, "colore 2"),
                "ponte": get(row, idx_s, "ponte"),
                "prezzo acquisto": get(row, idx_s, "prezzo acquisto"),
                "prezzo di vendita": get(row, idx_s, "prezzo di vendita"),
            }
    fh_s.close()

    print("Lettura MOVIMENTI ACQUISTO (solo barcode rilevanti)...", file=sys.stderr)
    # Colonna "data" (AB) = data piu' recente di acquisto per codice a barre,
    # calcolata da MOVIMENTI ACQUISTO.txt (solo tipo operazione = ACQ).
    data_ultimo_acquisto = {}
    idx_m, reader_m, fh_m = iter_tsv(p("movimenti"))
    for row in reader_m:
        bc = get(row, idx_m, "codice a barre")
        if bc not in barcodes_rilevanti:
            continue
        if get(row, idx_m, "tipo operazione") != "ACQ":
            continue
        d = parse_it_date(get(row, idx_m, "data"))
        if not d:
            continue
        if bc not in data_ultimo_acquisto or d > data_ultimo_acquisto[bc]:
            data_ultimo_acquisto[bc] = d
    fh_m.close()

    print("Lettura VENDITE (file unico multi-anno, solo tipo operazione VEN)...", file=sys.stderr)
    import openpyxl as _openpyxl

    vendite_qta = defaultdict(int)
    vendite_qta_cumulato = defaultdict(int)
    ultima_vendita = {}

    anno_corrente = year
    anno_precedente = year - 1
    anno_meno2 = year - 2

    wb_v = _openpyxl.load_workbook(p("vendite"), read_only=True, data_only=True)
    ws_v = wb_v.active
    rows_v = ws_v.iter_rows(values_only=True)
    header_v = next(rows_v)
    idx_v = {name: i for i, name in enumerate(header_v)}

    def vget(row, name):
        i = idx_v.get(name)
        if i is None or i >= len(row):
            return None
        return row[i]

    n_vendite = 0
    for row in rows_v:
        if vget(row, "Tipo operazione") != "VEN":
            continue
        bc = vget(row, "Codice a barre")
        fil = vget(row, "Filiale")
        if not bc or not fil:
            continue
        d = vget(row, "Data")
        if d is None:
            continue
        d = d.date() if hasattr(d, "date") else d
        raw_qta = vget(row, "Quantità")
        try:
            qta = int(raw_qta or 0)
        except (TypeError, ValueError):
            qta = 0
        n_vendite += 1
        key = (bc, fil, d.year)
        vendite_qta[key] += qta
        cur = ultima_vendita.get((bc, fil))
        if cur is None or d > cur:
            ultima_vendita[(bc, fil)] = d
        if d.year == anno_precedente and (d.month, d.day) <= (today.month, today.day):
            vendite_qta_cumulato[(bc, fil)] += qta
    wb_v.close()
    print(f"  righe vendita (VEN) lette: {n_vendite}", file=sys.stderr)

    col_vendita_meno2 = f"Vendita {anno_meno2}"
    col_vendita_prec = f"Vendita {anno_precedente}"
    col_vendita_prec_cum = f"Vendita {anno_precedente} CUMULATO"
    col_vendita_corr = f"Vendita {anno_corrente}"

    ref_fornitore2 = load_ref_csv(os.path.join(ref_dir, "ref_fornitore2.csv"), "fornitore")
    ref_articolo = load_ref_csv(os.path.join(ref_dir, "ref_articolo_manuale.csv"), "Codice a Barre")
    marchi_attivi_sole = load_brand_set(os.path.join(ref_dir, "ref_marchi_attivi_sole.csv"))
    marchi_attivi_vista = load_brand_set(os.path.join(ref_dir, "ref_marchi_attivi_vista.csv"))

    new_fornitori = [0]
    new_articoli = [0]

    def get_fornitore2(fornitore, marchio):
        # Regola fissa (non da tabella manuale): il marchio CARTIER e'
        # sempre distribuito da Kering Eyewear.
        if marchio == "CARTIER":
            return "KERING EYEWEAR"
        row = ref_fornitore2.get(fornitore)
        if row is None:
            ref_fornitore2[fornitore] = {"fornitore": fornitore, "Fornitore 2": ""}
            new_fornitori[0] += 1
            return ""
        return row.get("Fornitore 2", "")

    def get_marchi_attivi(marchio, tipo_lenti):
        m = (marchio or "").strip().upper()
        t = (tipo_lenti or "").strip().lower()
        if t == "sole" and m in marchi_attivi_sole:
            return "SI"
        if t == "vista" and m in marchi_attivi_vista:
            return "SI"
        return ""

    def get_articolo_manuale(barcode, marchio, modello):
        row = ref_articolo.get(barcode)
        if row is None:
            row = {
                "Codice a Barre": barcode, "Marchio": marchio, "modello": modello,
                "Top": "", "POSIZIONE GRIGLIA": "", "Occhiali con CLIP": "",
                "Personale": "", "scorta minima": "",
            }
            ref_articolo[barcode] = row
            new_articoli[0] += 1
        return row

    print("Costruzione righe...", file=sys.stderr)
    out_rows = []
    for grow in giac_rows:
        fil_code = get(grow, idx_g, "filiale")
        barcode = get(grow, idx_g, "codice a barre")
        if not barcode:
            continue

        li = listini_by_barcode.get(barcode, {})
        so = sottoscorta_by_barcode.get(barcode, {})

        modello = get(grow, idx_g, "modello")
        colore = get(grow, idx_g, "colore")
        calibro = get(grow, idx_g, "calibro")
        categoria = get(grow, idx_g, "categoria filtro")
        marchio = get(grow, idx_g, "Marchio")
        fornitore = get(grow, idx_g, "fornitore")
        tipo_lenti = get(grow, idx_g, "Tipo Lenti")

        # Ray-Ban: rimuovi SOLE/VISTA/SUN/OPTICAL dal modello PRIMA di
        # costruire le colonne G/H/I (che usano "modello").
        modello = clean_modello_rayban(marchio, modello)

        parts_cat = [x for x in (modello, colore, calibro, categoria) if x]
        parts_no_cat = [x for x in (modello, colore, calibro) if x]
        parts_col_cal = [x for x in (colore, calibro) if x]

        # Colonna G, costruita subito per poterla usare come base per la
        # colonna AH (Occhiali con CLIP).
        col_g_value = "; ".join(parts_cat)

        prezzo_acq_listino = parse_it_number(li.get("prezzo acquisto"))
        prezzo_ven_listino = parse_it_number(li.get("prezzo di vendita"))
        # AJ/AK (e i calcoli a valle Sconto Acquisto/Sconto Vendita/Fattore di
        # RICARICO che li usano come base) vengono presi dalla GIACENZA, non
        # piu' dalla sottoscorta.
        prezzo_acq_scheda = parse_it_number(get(grow, idx_g, "prezzo acquisto"))
        prezzo_ven_scheda = parse_it_number(get(grow, idx_g, "prezzo di vendita"))

        # AN/AO: valori interi arrotondati (nessun decimale).
        sconto_acq = None
        if prezzo_acq_listino:
            base = prezzo_acq_scheda if prezzo_acq_scheda is not None else prezzo_acq_listino
            sconto_acq = round(100 * (1 - base / prezzo_acq_listino))
        sconto_ven = None
        if prezzo_ven_listino:
            base = prezzo_ven_scheda if prezzo_ven_scheda is not None else prezzo_ven_listino
            sconto_ven = round(100 * (1 - base / prezzo_ven_listino))
        # AP: un solo decimale.
        ricarico = None
        if prezzo_acq_scheda:
            ricarico = round((prezzo_ven_scheda or 0) / prezzo_acq_scheda, 1)

        def to_int(s):
            try:
                return int(float((s or "0").replace(",", ".")))
            except ValueError:
                return 0

        q_mag = to_int(get(grow, idx_g, "quantita magazzino"))
        q_prenot = to_int(get(grow, idx_g, "quantita prenotata"))
        q_arrivo = to_int(get(grow, idx_g, "quantita in arrivo"))
        q_disp = max(0, q_mag - q_prenot)

        v_corrente = vendite_qta.get((barcode, fil_code, anno_corrente)) or 0
        v_precedente = vendite_qta.get((barcode, fil_code, anno_precedente)) or 0

        # Filtro: escludi le combinazioni articolo+negozio con giacenza zero
        # E nessuna vendita negli ultimi due anni (anno corrente + anno precedente).
        if q_mag == 0 and v_corrente == 0 and v_precedente == 0:
            continue

        manual = get_articolo_manuale(barcode, marchio, modello)

        row = {
            "fornitore": fornitore,
            "Fornitore 2": get_fornitore2(fornitore, marchio),
            "Marchio": marchio,
            "modello": modello,
            "Modello CARTIER": modello if marchio == "CARTIER" else "",
            "SKU": li.get("SKU") or so.get("SKU") or "",
            "Modello + Colore + Calibro + CAT": col_g_value,
            "Modello + Colore + Calibro": "; ".join(parts_no_cat),
            "Colore + Calibro": "; ".join(parts_col_cal),
            "colore": colore,
            "colore 2": li.get("colore 2") or so.get("colore 2") or "",
            "calibro": calibro,
            "ponte": li.get("ponte") or so.get("ponte") or "",
            "materiale": get(grow, idx_g, "materiale"),
            "utente": get(grow, idx_g, "utente"),
            "Tipo Lenti": tipo_lenti,
            "Filiale": FILIALI[fil_code],
            "Filiale BIS": FILIALI[fil_code],
            "quantita magazzino": q_mag,
            "Quantita Disponibile": q_disp,
            "quantita in arrivo": q_arrivo,
            "scorta minima": manual.get("scorta minima", ""),
            col_vendita_meno2: (vendite_qta.get((barcode, fil_code, anno_meno2)) or ""),
            col_vendita_prec: v_precedente or "",
            col_vendita_prec_cum: vendite_qta_cumulato.get((barcode, fil_code)) or "",
            col_vendita_corr: v_corrente or "",
            "data ult acquisto": parse_it_date(get(grow, idx_g, "data ult. acquisto")),
            "data": data_ultimo_acquisto.get(barcode),
            "data ult vendita": ultima_vendita.get((barcode, fil_code)),
            "Categoria Filtro": categoria,
            "POSIZIONE GRIGLIA": manual.get("POSIZIONE GRIGLIA", ""),
            "Top": manual.get("Top", ""),
            "Marchi Attivi": get_marchi_attivi(marchio, tipo_lenti),
            "Occhiali con CLIP": get_occhiali_clip(tipo_lenti, col_g_value),
            "Personale": manual.get("Personale", ""),
            "Prezzo Di Acquisto Scheda Scontato": prezzo_acq_scheda,
            "Prezzo Di Vendita Scheda Scontato": prezzo_ven_scheda,
            "Prezzo Acquisto Listino Intero": prezzo_acq_listino,
            "Prezzo Vendita Listino Intero": prezzo_ven_listino,
            "Sconto Acquisto": sconto_acq,
            "Sconto Vendita": sconto_ven,
            "Fattore di RICARICO (su prezzi scontati)": ricarico,
            "Codice a Barre": barcode,
            "ID_GIACENZA": f"{barcode}{fil_code}",
            "Filiale Focus": fil_code,
        }
        out_rows.append(row)

    columns = [c for c in OUTPUT_COLUMNS if c is not None]
    idx = columns.index("scorta minima") + 1
    columns[idx:idx] = [col_vendita_meno2, col_vendita_prec, col_vendita_prec_cum, col_vendita_corr]

    os.makedirs(ref_dir, exist_ok=True)
    save_ref_csv(os.path.join(ref_dir, "ref_fornitore2.csv"), ["fornitore", "Fornitore 2"], "fornitore", ref_fornitore2)
    save_ref_csv(
        os.path.join(ref_dir, "ref_articolo_manuale.csv"),
        ["Codice a Barre", "Marchio", "modello", "Top", "POSIZIONE GRIGLIA", "Occhiali con CLIP", "Personale", "scorta minima"],
        "Codice a Barre", ref_articolo,
    )

    print(
        f"Fatto. Righe: {len(out_rows)}. Nuovi fornitori: {new_fornitori[0]}, "
        f"nuovi articoli: {new_articoli[0]} "
        f"(aggiunti vuoti nelle tabelle di riferimento, da compilare). "
        f"Marchi attivi: {len(marchi_attivi_sole)} sole, {len(marchi_attivi_vista)} vista.",
        file=sys.stderr,
    )
    return columns, out_rows


def write_xlsx(columns, rows, output_path):
    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Base")
    ws.append(columns)
    for row in rows:
        ws.append([row.get(c) for c in columns])
    wb.save(output_path)


def main():
    ap = argparse.ArgumentParser(description="Genera il file Base MONTATURE dai file sorgente.")
    ap.add_argument("--input-dir", required=True, help="Cartella con i file BASE/ e RICERCHE BASE/")
    ap.add_argument("--ref-dir", required=True, help="Cartella con le tabelle di riferimento (CSV)")
    ap.add_argument("--output", required=True, help="Percorso del file xlsx da generare")
    args = ap.parse_args()

    columns, rows = build(args.input_dir, args.ref_dir)
    write_xlsx(columns, rows, args.output)
    print(f"Salvato: {args.output}", file=sys.stderr)


if __name__ == "__main__":
    main()
